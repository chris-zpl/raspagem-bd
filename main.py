# Imports para automatizar usando o Selenium
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import WebDriverException, NoSuchElementException
from selenium.webdriver.common.by import By
from time import sleep
from openpyxl import Workbook
from openpyxl.styles import Alignment
from params import usuarios, site_exec, cores, estilo_sheet  # Meu arquivo
import os

class Scrapy:
    def __init__(self):
        opcao = Options()
        #opcao.add_argument('--window-size=1920,1080')
        opcao.add_argument('--disable-notifications')
        opcao.add_argument('--disable-extensions')
        opcao.add_argument('--disable-gpu')
        opcao.add_argument('--start-maximized')
        servico = Service(ChromeDriverManager().install())
        self.navegador = webdriver.Chrome(options=opcao, service=servico)
    
    def iniciar(self):
        user = usuarios()
        for usuario, senha in user.items():
            try:
                print(cores(cor='cyan', mensagem='\n## AUTOMAÇÃO WEB SCRAPING ##'))
                self.login(usuario, senha)
            except WebDriverException as e:
                print(cores(cor='vermelho', mensagem='\nErro inesperado. Encerrando execução.'))
                print(e) # Apenas para Debug
                sleep(5)
    
    def login(self, usuario, senha):
        self.navegador.get(site_exec())
        print(cores(cor='marrom', mensagem='\n>> Efetuando Login...'))
        sleep(2)

        # Envia as credenciais para login
        self.navegador.find_element(
            By.NAME, 'loginUsername').send_keys(usuario)
        sleep(2)
        self.navegador.find_element(
            By.ID, 'loginPassword').send_keys(senha)
        sleep(2)
        self.navegador.find_element(By.NAME, 'loginSubmit').click()
        sleep(2)
        # Realiza um try para verifica se, ao digitar o usuário e a senha, ele irá avançar para a página de autenticação.
        # Caso não aconteça, fecha o navegador e informa o problema para ser corrigido.
        try:
            self.navegador.find_element(By.XPATH, '/html/body/table/tbody/tr/td[1]/a[1]/span[2]')
            print(cores(cor='verde', mensagem='>> Login efetuado\n'))
            # Faz a execução da função
            self.execucao()
        except NoSuchElementException:
            print(cores(cor='vermelho', mensagem='O nome de usuário ou senha está incorreto \nCorrija antes de iniciá-lo novamente\n'))
            print(cores(cor='branco', mensagem='Finalizando execução.'))
            sleep(5)
            self.navegador.quit()

    def execucao(self):
        try:
            print(cores(cor='branco', mensagem='Aguardando o início da execução\n'))
            self.raspagem_dados_professores()
            print(cores(cor='branco', mensagem='\nAguardando o início da próxima execução'))
            self.raspagem_dados_esp_fisico()
            self.criar_planilha()
            print(cores(cor='branco', mensagem='\nFinalizando a execução'))
            sleep(2)
            self.navegador.quit()  
        except WebDriverException as e:
            print(cores(cor='vermelho', mensagem='Algo deu errado durante a execução\n'))
            print(e) # Apenas para Debug
            print(cores(cor='branco', mensagem='Finalizando execução.'))
            sleep(3)
            self.navegador.quit()

    def acesso_tabela_bd(self, tabela):
        self.navegador.find_element(By.LINK_TEXT, 'feng_ementas_20102').click()
        sleep(2)
        self.navegador.find_element(By.LINK_TEXT, '2023-2').click()
        sleep(2)
        self.navegador.find_element(By.LINK_TEXT, tabela).click()
        sleep(2)
        self.navegador.find_element(By.LINK_TEXT, 'Navegar').click()
        sleep(2)

    def acesso_info_bd(self, linha, coluna):
        try:
            resultado = int(self.navegador.find_element(By.XPATH, f'//*[@id="data"]/tbody/tr[{linha}]/td[{coluna}]').text)
        except:
            resultado = self.navegador.find_element(By.XPATH, f'//*[@id="data"]/tbody/tr[{linha}]/td[{coluna}]').text

        return resultado

    def raspagem_dados_professores(self):
        # Listas vazias para receber os valores
        self.lista_cod_prof = []
        self.lista_nomes_prof = []
        self.lista_unidade_prof = []
        self.lista_matricula_prof = []
        self.lista_email_prof = []
        self.lista_espaco_fisico_prof = []

        # Ao realizar o login, clica nos caminhos para ir até a tabela dos professores
        self.acesso_tabela_bd(tabela='professor')

        print(cores(cor='cyan', mensagem='Iniciando Web Scraping dos Professores\n'))

        # Um click na coluna nome para dar um sorted em ASC
        self.navegador.find_element(By.XPATH, f'//*[@id="data"]/tbody/tr[1]/th[3]').click()

        tr = 2
        pg = 1
        while True:
            # Conta a quantidade de elementos na tabela
            elementos = self.navegador.find_elements(By.XPATH, f'//*[@id="data"]/tbody//td[3]')
            td_cont = len(elementos)
            # Faz a raspagem dos dados e adiciona para as listas
            for i in range(td_cont):
                # Lista com os códigos
                lista_cod = self.acesso_info_bd(linha=tr, coluna='3')
                self.lista_cod_prof.append(lista_cod)

                # Lista com os nomes
                lista_nomes = self.acesso_info_bd(linha=tr, coluna='4')
                self.lista_nomes_prof.append(lista_nomes)

                # Lista com a unidade
                lista_unidade = self.acesso_info_bd(linha=tr, coluna='6')
                self.lista_unidade_prof.append(lista_unidade)

                # Lista com a matrícula
                lista_matricula = self.acesso_info_bd(linha=tr, coluna='8')
                self.lista_matricula_prof.append(lista_matricula)

                # Lista com o email
                lista_email = self.acesso_info_bd(linha=tr, coluna='11')
                self.lista_email_prof.append(lista_email)

                # Lista com o espaço físico
                lista_espaco_fisico = self.acesso_info_bd(linha=tr, coluna='14')
                self.lista_espaco_fisico_prof.append(lista_espaco_fisico)

                tr += 1
            pg += 1
            try:
                self.navegador.find_element(By.LINK_TEXT, 'Próximo >').click()
                print(cores(cor='marrom', mensagem=f'>> Avançou para a página {pg}'))
                tr = 2
            except NoSuchElementException:
                print(cores(cor='marrom', mensagem=f'\nNão há mais páginas!'))
                print(cores(cor='verde', mensagem='Escaneamento concluído'))
                break
        
        self.navegador.find_element(By.LINK_TEXT, 'feng_ementas_20102').click()

    def raspagem_dados_esp_fisico(self):
        # Listas vazias para receber os valores
        self.lista_cod_salas = []
        self.lista_espaco_fisico = []
        self.lista_nome_salas = []
        self.lista_capacidade_salas = []
        # Ao realizar o login, clica nos caminhos para ir até a tabela dos professores
        self.acesso_tabela_bd(tabela='espaco_fisico')

        print(cores(cor='cyan', mensagem='\nIniciando Web Scraping dos Espaços Físicos\n'))

        # Um click na coluna código para dar um sorted em ASC
        self.navegador.find_element(By.XPATH, f'//*[@id="data"]/tbody/tr[1]/th[2]').click()

        tr = 2
        pg = 1
        while True:
            # Conta a quantidade de elementos na tabela
            elementos = self.navegador.find_elements(By.XPATH, f'//*[@id="data"]/tbody//td[3]')
            td_cont = len(elementos)
            # Faz a raspagem dos dados e adiciona para as listas
            for i in range(td_cont):
                # Lista com os códigos
                lista_cod = self.acesso_info_bd(linha=tr, coluna='3')
                self.lista_cod_salas.append(lista_cod)

                # Lista com as salas
                predio = str(self.acesso_info_bd(linha=tr, coluna='4'))
                bloco = str(self.acesso_info_bd(linha=tr, coluna='5'))
                sala = str(self.acesso_info_bd(linha=tr, coluna='6'))
                lista_salas = predio + '/' + bloco + '/' + sala
                self.lista_espaco_fisico.append(lista_salas)

                # Lista com os nomes das salas
                lista_nomes = self.acesso_info_bd(linha=tr, coluna='7')
                self.lista_nome_salas.append(lista_nomes)

                # Lista com a capacidade de cada sala
                lista_capacidade = self.acesso_info_bd(linha=tr, coluna='15')
                self.lista_capacidade_salas.append(lista_capacidade)

                tr += 1
            pg += 1
            try:
                self.navegador.find_element(By.LINK_TEXT, 'Próximo >').click()
                print(cores(cor='marrom', mensagem=f'>> Avançou para a página {pg}'))
                tr = 2
            except NoSuchElementException:
                print(cores(cor='marrom', mensagem=f'\nNão há mais páginas!'))
                print(cores(cor='verde', mensagem='Escaneamento concluído'))
                break

        self.navegador.find_element(By.LINK_TEXT, 'feng_ementas_20102').click()

    def criar_planilha(self):
        print(cores('cyan', '\n>> Gerando planilha em Excel'))
        sleep(1)

        # A planilha dos professores
        try:
            if not self.lista_cod_prof == '':
                index = 2
                wb = Workbook()
                ws = wb.active
                ws.title = 'Código dos Professores'
                ws.sheet_properties.tabColor = 'FFC000'
                ws['A1'] = 'Código'
                estilo_sheet(ws,'A1')
                ws['B1'] = 'Nome'
                estilo_sheet(ws,'B1')
                ws['C1'] = 'Unidade'
                estilo_sheet(ws,'C1')
                ws['D1'] = 'Matrícula'
                estilo_sheet(ws,'D1')
                ws['E1'] = 'E-mail'
                estilo_sheet(ws,'E1')
                ws['F1'] = 'Espaço Físico'
                estilo_sheet(ws,'F1')
                tamanho_max_nome = 0
                tamanho_max_unidade = 0
                tamanho_max_email = 0
                for cod, nome, uni, matric, email, espacFis in zip(self.lista_cod_prof, self.lista_nomes_prof, self.lista_unidade_prof, self.lista_matricula_prof, self.lista_email_prof, self.lista_espaco_fisico_prof):
                    ws.cell(column=1, row=index, value=cod).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(column=2, row=index, value=nome).alignment = Alignment(wrap_text=True, vertical='center')
                    ws.cell(column=3, row=index, value=uni).alignment = Alignment(horizontal='center',wrap_text=True, vertical='center')
                    ws.cell(column=4, row=index, value=matric).alignment = Alignment(horizontal='center',wrap_text=True, vertical='center')
                    ws.cell(column=5, row=index, value=email).alignment = Alignment(horizontal='center',wrap_text=True, vertical='center')
                    ws.cell(column=6, row=index, value=espacFis).alignment = Alignment(horizontal='center',wrap_text=True, vertical='center')
                    # Calcula o tamanho da cell e adapta o tamanho
                    if len(nome) > tamanho_max_nome:
                        tamanho_max_nome = len(nome)
                    ws.column_dimensions['B'].width = tamanho_max_nome + 1

                    if len(uni) > tamanho_max_unidade:
                        tamanho_max_unidade = len(uni)
                    ws.column_dimensions['C'].width = tamanho_max_unidade + 1


                    if len(email) > tamanho_max_email:
                        tamanho_max_email = len(email)
                    ws.column_dimensions['E'].width = tamanho_max_email + 1

                    ws.column_dimensions['F'].width = 12.20

                    index += 1
                print(cores(cor='verde', mensagem=f'Planilha "{ws.title}" criada com sucesso'))
        except:
            print(cores(cor='vermelho', mensagem=f'\nNão foi possível criar uma planilha relacionada aos professores'))
            sleep(2)
        
        # Planilha dos Espaços Físicos
        try:
            if not self.lista_cod_salas == '':
                index = 2
                ws2 = wb.create_sheet()
                ws2.title = 'Código dos Espaços Físicos'
                ws2.sheet_properties.tabColor = '538ED5'
                ws2['A1'] = 'Código'
                estilo_sheet(ws2,'A1')
                ws2['B1'] = 'Salas'
                estilo_sheet(ws2,'B1')
                ws2['C1'] = 'Nome'
                estilo_sheet(ws2,'C1')
                ws2['D1'] = 'Capacidade'
                estilo_sheet(ws2,'D1')
                tamanho_max_salas = 0
                tamanho_max_nomes = 0
                for cod, salas, nome, capacidade in zip(self.lista_cod_salas, self.lista_espaco_fisico, self.lista_nome_salas, self.lista_capacidade_salas):
                    ws2.cell(column=1, row=index, value=cod).alignment = Alignment(horizontal='center', vertical='center')
                    ws2.cell(column=2, row=index, value=salas).alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
                    ws2.cell(column=3, row=index, value=nome).alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
                    ws2.cell(column=4, row=index, value=capacidade).alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
                    # Calcula o tamanho da cell e adapta o tamanho
                    if len(salas) > tamanho_max_salas:
                        tamanho_max_salas = len(salas)
                    ws2.column_dimensions['B'].width = tamanho_max_salas + 1
                    if len(nome) > tamanho_max_nomes:
                        tamanho_max_nomes = len(nome)
                    ws2.column_dimensions['C'].width = tamanho_max_nomes + 1

                    ws2.column_dimensions['D'].width = 11
                    index += 1

                print(cores(cor='verde', mensagem=f'Planilha "{ws2.title}" criada com sucesso'))
        except:
            print(cores(cor='vermelho', mensagem=f'\nNão foi possível criar uma planilha relacionada ao espaço físico'))
            sleep(2)

        # Para salvar o documento no diretório "Resultado"
        script_dir = os.path.dirname(__file__)
        caminho_abs = os.path.join(script_dir, 'resultado', 'planilha_bd.xlsx')
        wb.save(caminho_abs)
        print(cores(cor='verde', mensagem='\nArquivo gerado com sucesso'))
        sleep(2)

# Cria uma variável chamada raspagem e inicia a função
if __name__ == '__main__':
    raspagem = Scrapy()
    raspagem.iniciar()
