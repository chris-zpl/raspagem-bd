import os
from openpyxl.styles import Alignment, PatternFill

script_dir = os.path.dirname(__file__)
abs_arq_users = os.path.join(script_dir, 'infos', 'usuarios.csv')
abs_arq_path_site = os.path.join(script_dir, 'infos', 'site_exec.txt')


def usuarios():
    '''Função para buscar usuários no arquivo CSV'''
    dados_users = {}
    with open(abs_arq_users, 'r') as arquivo:
        # Ignorar a primeira linha
        arquivo.readline()
        for linha in arquivo:
            valores = linha[:-1].split(',')
            dados_users[valores[0]] = valores[1]
    return dados_users


def site_exec():
    '''Função para ler o site que será executado'''
    with open(abs_arq_path_site, 'r') as arquivo:
        dados_exec = arquivo.read()

    return dados_exec


def cores(cor, mensagem):
    '''Parâmetro de cores'''

    match cor:
        case 'vermelho':
            vermelho = (f'\033[0;31m{mensagem}\033[0m')
            return vermelho
        case 'verde':
            verde = (f'\033[0;32m{mensagem}\033[0m')
            return verde
        case 'marrom':
            marrom = (f'\033[0;33m{mensagem}\033[0m')
            return marrom
        case 'purple':
            purple = (f'\033[0;35m{mensagem}\033[0m')
            return purple
        case 'cyan':
            cyan = (f'\033[0;36m{mensagem}\033[0m')
            return cyan
        case 'branco':
            branco = (f'\033[1;37m{mensagem}\033[0m')
            return branco


def estilo_sheet(ws, cell):
    ws[cell].alignment = Alignment(horizontal='center', vertical='center')
    ws[cell].fill = PatternFill(fgColor="C9C9C9", fill_type='solid')


